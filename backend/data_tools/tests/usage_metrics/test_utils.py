import io
from datetime import datetime, timedelta, timezone
from unittest.mock import MagicMock

import pytest
from openpyxl import load_workbook
from sqlalchemy import text

from data_tools.src.usage_metrics.utils import (
    METRIC_COLUMNS,
    aggregate_events,
    aggregate_user_sign_ins,
    build_workbook,
    is_deactivating_update,
    parse_lookback_days,
    resolve_actor_id,
    run_usage_metrics,
)
from models import Division, OpsEvent, OpsEventStatus, OpsEventType, Role, User, UserStatus

# Wide window for the counting tests; a couple of tests override it to exercise the boundary.
LOOKBACK_DAYS = 30

# ---------------------------------------------------------------------------
# Pure-unit tests for the attribution helpers (no DB).
# ---------------------------------------------------------------------------


def _event(event_type, *, created_by=None, event_details=None, event_status=OpsEventStatus.SUCCESS):
    ev = OpsEvent(
        event_type=event_type,
        event_status=event_status,
        created_by=created_by,
        event_details=event_details,
    )
    return ev


def test_resolve_actor_id_uses_created_by_for_non_login_events():
    ev = _event(OpsEventType.GET_AGREEMENT, created_by=42)
    assert resolve_actor_id(ev) == 42


def test_resolve_actor_id_reads_login_actor_from_event_details():
    ev = _event(
        OpsEventType.LOGIN_ATTEMPT,
        created_by=None,
        event_details={"user": {"id": 7}},
    )
    assert resolve_actor_id(ev) == 7


def test_resolve_actor_id_returns_none_for_failed_login_without_user_key():
    # FAILED login rows have no "user" key in event_details.
    ev = _event(
        OpsEventType.LOGIN_ATTEMPT,
        created_by=None,
        event_details={"error_message": "bad token"},
        event_status=OpsEventStatus.FAILED,
    )
    assert resolve_actor_id(ev) is None


def test_resolve_actor_id_tolerates_missing_event_details():
    ev = _event(OpsEventType.LOGIN_ATTEMPT, created_by=None, event_details=None)
    assert resolve_actor_id(ev) is None


def test_is_deactivating_update_true_for_status_inactive():
    ev = _event(OpsEventType.UPDATE_USER, created_by=1, event_details={"request.json": {"status": "INACTIVE"}})
    assert is_deactivating_update(ev) is True


def test_is_deactivating_update_true_for_status_locked():
    ev = _event(OpsEventType.UPDATE_USER, created_by=1, event_details={"request.json": {"status": "LOCKED"}})
    assert is_deactivating_update(ev) is True


def test_is_deactivating_update_false_for_other_status_change():
    ev = _event(OpsEventType.UPDATE_USER, created_by=1, event_details={"request.json": {"status": "ACTIVE"}})
    assert is_deactivating_update(ev) is False


def test_is_deactivating_update_false_for_non_update_user_event():
    ev = _event(OpsEventType.CREATE_USER, created_by=1, event_details={"request.json": {"status": "INACTIVE"}})
    assert is_deactivating_update(ev) is False


def test_is_deactivating_update_tolerates_missing_payload():
    ev = _event(OpsEventType.UPDATE_USER, created_by=1, event_details={})
    assert is_deactivating_update(ev) is False


def test_is_deactivating_update_true_for_automated_top_level_status():
    # The disable_users job emits a top-level "status" (no request.json wrapper).
    ev = _event(
        OpsEventType.UPDATE_USER,
        created_by=1,
        event_details={
            "user_id": 5,
            "status": UserStatus.INACTIVE.name,
            "message": "User deactivated via automated process.",
        },
    )
    assert is_deactivating_update(ev) is True


def test_is_deactivating_update_false_for_automated_non_deactivating_status():
    ev = _event(
        OpsEventType.UPDATE_USER,
        created_by=1,
        event_details={"user_id": 5, "status": UserStatus.ACTIVE.name},
    )
    assert is_deactivating_update(ev) is False


def test_is_deactivating_update_matches_userstatus_enum_names():
    # The status literals are derived from the UserStatus enum, not hardcoded strings.
    for status in (UserStatus.INACTIVE, UserStatus.LOCKED):
        ev = _event(OpsEventType.UPDATE_USER, created_by=1, event_details={"request.json": {"status": status.name}})
        assert is_deactivating_update(ev) is True


def test_parse_lookback_days_valid():
    assert parse_lookback_days("7") == 7


def test_parse_lookback_days_invalid_raises():
    with pytest.raises(ValueError):
        parse_lookback_days("not-a-number")


@pytest.mark.parametrize("value", ["0", "-1"])
def test_parse_lookback_days_non_positive_raises(value):
    # A non-positive window would produce an empty report with no error; fail fast instead.
    with pytest.raises(ValueError):
        parse_lookback_days(value)


# ---------------------------------------------------------------------------
# Integration tests against a real DB (loaded_db, SAVEPOINT-isolated).
# ---------------------------------------------------------------------------


@pytest.fixture()
def seeded_db(loaded_db):
    """Seed a division, role, users, and ops_event rows with explicit created_on.

    The request-time OpsEventHandler does NOT run in tests, so created_by / event_details /
    created_on must be set explicitly on each seeded row. Timestamps are relative to now (naive
    UTC) so the reporting-window filter includes them regardless of the calendar date the test
    runs on.
    """
    division = Division(id=900, name="Test Division", abbreviation="TD")
    loaded_db.merge(division)

    role = Role(id=900, name="analyst")
    loaded_db.merge(role)
    loaded_db.commit()

    role = loaded_db.get(Role, 900)
    user_a = User(
        id=9001,
        email="a@example.com",
        first_name="Ada",
        last_name="Aardvark",
        division=900,
        status=UserStatus.ACTIVE,
        roles=[role],
    )
    user_b = User(
        id=9002,
        email="b@example.com",
        first_name="Ben",
        last_name="Bison",
        division=900,
        status=UserStatus.ACTIVE,
        roles=[role],
    )
    loaded_db.add_all([user_a, user_b])
    loaded_db.commit()

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    day1 = now - timedelta(days=3)
    day2 = now - timedelta(days=2)

    events = [
        # Two distinct viewers on day1 -> active_users = 2, agreements_viewed = 2.
        _event(OpsEventType.GET_AGREEMENT, created_by=9001),
        _event(OpsEventType.GET_AGREEMENT, created_by=9002),
        # Successful login on day1 -> logins = 1 (actor from event_details, created_by NULL).
        _event(OpsEventType.LOGIN_ATTEMPT, created_by=None, event_details={"user": {"id": 9001}}),
        # Failed login must be excluded (FAILED status filtered out at query time).
        _event(
            OpsEventType.LOGIN_ATTEMPT,
            created_by=None,
            event_details={"error_message": "nope"},
            event_status=OpsEventStatus.FAILED,
        ),
        # FAILED create must be excluded from blis_created (regression: previously counted).
        _event(OpsEventType.CREATE_BLI, created_by=9001, event_status=OpsEventStatus.FAILED),
        # BLI create and project create on day1 (SUCCESS).
        _event(OpsEventType.CREATE_BLI, created_by=9001),
        _event(OpsEventType.CREATE_PROJECT, created_by=9001),
        # Deactivation via UPDATE_USER on day2.
        _event(
            OpsEventType.UPDATE_USER,
            created_by=9001,
            event_details={"request.json": {"status": UserStatus.INACTIVE.name}},
        ),
    ]
    whens = [day1, day1, day1, day1, day1, day1, day1, day2]
    for ev, when in zip(events, whens, strict=True):
        ev.created_on = when
        loaded_db.add(ev)
    loaded_db.commit()

    yield loaded_db, day1.date().isoformat(), day2.date().isoformat()

    loaded_db.execute(text("DELETE FROM ops_event"))
    loaded_db.execute(text("DELETE FROM ops_event_version"))
    loaded_db.execute(text("DELETE FROM user_role"))
    loaded_db.execute(text("DELETE FROM user_role_version"))
    loaded_db.commit()


def test_aggregate_events_counts_by_day_division_role(seeded_db):
    db, day1_iso, day2_iso = seeded_db
    counts = aggregate_events(db, LOOKBACK_DAYS)

    day1_key = (day1_iso, "Test Division", "analyst")
    assert day1_key in counts
    day1 = counts[day1_key]
    assert day1["active_users"] == 2
    assert day1["agreements_viewed"] == 2
    assert day1["logins"] == 1  # failed login excluded
    assert day1["blis_created"] == 1  # FAILED create excluded, only the SUCCESS one counted
    assert day1["projects_created"] == 1
    assert day1["deactivated_users"] == 0

    day2_key = (day2_iso, "Test Division", "analyst")
    assert counts[day2_key]["deactivated_users"] == 1


def test_failed_events_excluded(seeded_db):
    db, _, _ = seeded_db
    counts = aggregate_events(db, LOOKBACK_DAYS)
    # Both a FAILED login and a FAILED CREATE_BLI were seeded; neither should be counted anywhere.
    assert sum(bucket["logins"] for bucket in counts.values()) == 1
    assert sum(bucket["blis_created"] for bucket in counts.values()) == 1


def test_events_outside_window_excluded(seeded_db):
    db, _, _ = seeded_db
    # All seeded events are 2-3 days old; a 1-day window excludes them all.
    counts = aggregate_events(db, 1)
    assert counts == {}


def test_utc_day_bucketing_boundary(loaded_db):
    """An event just before/after UTC midnight buckets to the correct UTC day."""
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    # Pick a recent day and place two events 20 minutes on either side of its UTC midnight.
    base = (now - timedelta(days=2)).date()
    before_midnight = datetime(base.year, base.month, base.day, 23, 50, 0)
    after_midnight = before_midnight + timedelta(minutes=20)  # next UTC day, 00:10

    ev1 = _event(OpsEventType.GET_AGREEMENT, created_by=None)
    ev1.created_on = before_midnight
    ev2 = _event(OpsEventType.GET_AGREEMENT, created_by=None)
    ev2.created_on = after_midnight
    loaded_db.add_all([ev1, ev2])
    loaded_db.commit()

    try:
        counts = aggregate_events(loaded_db, LOOKBACK_DAYS)
        dates = {date for (date, _div, _role) in counts.keys()}
        assert before_midnight.date().isoformat() in dates
        assert after_midnight.date().isoformat() in dates
        assert before_midnight.date() != after_midnight.date()
    finally:
        loaded_db.execute(text("DELETE FROM ops_event"))
        loaded_db.execute(text("DELETE FROM ops_event_version"))
        loaded_db.commit()


def test_run_usage_metrics_uploads_when_storage_configured(seeded_db, mocker):
    """When a storage account URL is configured, the .xlsx (dated + latest) is uploaded."""
    db, _, _ = seeded_db
    upload_mock = mocker.patch("data_tools.src.usage_metrics.utils.upload_blob")

    config = MagicMock()
    config.usage_metrics_storage_account_url = "https://acct.blob.core.windows.net"
    config.usage_metrics_container_name = "data"
    config.usage_metrics_report_prefix = "reports"
    config.usage_metrics_lookback_days = "30"

    conn = MagicMock()
    mocker.patch("data_tools.src.usage_metrics.utils.Session").return_value.__enter__.return_value = db

    run_usage_metrics(conn, config)

    # Two uploads: dated + latest .xlsx (no CSV).
    assert upload_mock.call_count == 2
    blob_names = {call.args[2] for call in upload_mock.call_args_list}
    assert "reports/usage-metrics-latest.xlsx" in blob_names
    assert any(name.startswith("reports/usage-metrics-") and name.endswith(".xlsx") for name in blob_names)
    # No CSV is produced any more.
    assert not any(name.endswith(".csv") for name in blob_names)

    # Every upload is an .xlsx carrying the spreadsheet MIME type.
    assert all(
        c.args[2].endswith(".xlsx")
        and c.kwargs.get("content_type") == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        for c in upload_mock.call_args_list
    )


# ---------------------------------------------------------------------------
# Per-user sign-in aggregation and workbook.
# ---------------------------------------------------------------------------


def test_aggregate_user_sign_ins_counts_per_user_not_per_role(seeded_db, loaded_db):
    """sign_in_count is keyed on the actor only, so a multi-role user is not double-counted."""
    db, _, _ = seeded_db

    # Give user_b (9002) a second role; the login count must stay per-user, not fan out per role.
    second_role = Role(id=901, name="reviewer")
    db.merge(second_role)
    db.commit()
    user_b = db.get(User, 9002)
    user_b.roles.append(db.get(Role, 901))
    db.commit()

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    # user_a (9001) already has one login on day1 from the fixture; add a later one on day2.
    later = now - timedelta(days=1)
    extra_a = _event(OpsEventType.LOGIN_ATTEMPT, created_by=None, event_details={"user": {"id": 9001}})
    extra_a.created_on = later
    # user_b (9002, two roles) has a single login.
    login_b = _event(OpsEventType.LOGIN_ATTEMPT, created_by=None, event_details={"user": {"id": 9002}})
    login_b.created_on = now - timedelta(days=2)
    db.add_all([extra_a, login_b])
    db.commit()

    rows = aggregate_user_sign_ins(db, LOOKBACK_DAYS)
    by_email = {r["email"]: r for r in rows}

    assert by_email["a@example.com"]["sign_in_count"] == 2
    assert by_email["a@example.com"]["name"] == "Ada Aardvark"
    assert by_email["a@example.com"]["last_sign_in_utc"] == later.isoformat()

    # Two roles, but a single login -> count 1, and both roles joined in one column.
    assert by_email["b@example.com"]["sign_in_count"] == 1
    assert set(by_email["b@example.com"]["roles"].split(", ")) == {"analyst", "reviewer"}


def test_aggregate_user_sign_ins_excludes_failed_and_out_of_window(seeded_db):
    """Only SUCCESS logins inside the window are counted; a 1-day window excludes the seeded rows."""
    db, _, _ = seeded_db
    # The fixture seeds one SUCCESS login (9001) and one FAILED login, both ~3 days old.
    assert aggregate_user_sign_ins(db, LOOKBACK_DAYS) and all(
        r["sign_in_count"] >= 1 for r in aggregate_user_sign_ins(db, LOOKBACK_DAYS)
    )
    # The FAILED login is never counted (only user_a appears).
    emails = {r["email"] for r in aggregate_user_sign_ins(db, LOOKBACK_DAYS)}
    assert emails == {"a@example.com"}
    # A window that predates the seeded events yields no rows.
    assert aggregate_user_sign_ins(db, 1) == []


def test_aggregate_user_sign_ins_falls_back_to_event_details_for_unknown_user(seeded_db):
    """A login by a user id absent from the live table renders from the event_details snapshot."""
    db, _, _ = seeded_db
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    ghost = _event(
        OpsEventType.LOGIN_ATTEMPT,
        created_by=None,
        event_details={"user": {"id": 99999, "email": "ghost@example.com", "full_name": "Ghost User"}},
    )
    ghost.created_on = now - timedelta(days=1)
    db.add(ghost)
    db.commit()

    rows = aggregate_user_sign_ins(db, LOOKBACK_DAYS)
    ghost_row = next(r for r in rows if r["email"] == "ghost@example.com")
    assert ghost_row["name"] == "Ghost User"
    assert ghost_row["division"] == "UNKNOWN"
    assert ghost_row["roles"] == "UNKNOWN"
    assert ghost_row["sign_in_count"] == 1


def test_build_workbook_has_two_sheets_with_expected_columns():
    counts = {
        ("2026-07-06", "OD", "role_a"): {metric: 1 for metric in METRIC_COLUMNS},
    }
    user_rows = [
        {
            "name": "Ada Aardvark",
            "email": "a@example.com",
            "division": "OD",
            "roles": "analyst",
            "sign_in_count": 3,
            "last_sign_in_utc": "2026-07-06T12:00:00",
        }
    ]
    workbook_bytes = build_workbook(counts, user_rows)
    wb = load_workbook(io.BytesIO(workbook_bytes))

    assert wb.sheetnames == ["Aggregate", "Per-user"]

    agg = wb["Aggregate"]
    assert [c.value for c in agg[1]][:3] == ["date", "division", "role"]
    assert [c.value for c in agg[2]][:3] == ["2026-07-06", "OD", "role_a"]

    users = wb["Per-user"]
    assert [c.value for c in users[1]] == ["name", "email", "division", "roles", "sign_in_count", "last_sign_in_utc"]
    assert [c.value for c in users[2]] == ["Ada Aardvark", "a@example.com", "OD", "analyst", 3, "2026-07-06T12:00:00"]
