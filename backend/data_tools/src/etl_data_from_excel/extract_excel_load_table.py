import luigi
import openpyxl
from data_tools.src.etl_data_from_excel.models import AllBudgetCurrent, Base
from data_tools.src.etl_data_from_excel.utils import clean_rows
from pyexcel_io import save_data
from pyexcel_io.constants import DB_SQL
from pyexcel_io.database.common import SQLTableImportAdapter, SQLTableImporter
from sqlalchemy.future import create_engine
from sqlalchemy.orm import Session


class ExtractExcelLoadTable(luigi.Task):
    def output(self):
        ...

    def run(self):
        workbook = openpyxl.load_workbook(
            "data/REDACTED-FY22_Budget_Summary-10-12-22.xlsm",
            data_only=True,
            read_only=True,
            keep_vba=False,
            keep_links=False,
        )
        sheet = workbook["All Budget - Cur"]

        # header_row = [
        #     item.replace(" ", "_")
        #     for row in sheet.iter_rows(max_row=1, values_only=True)
        #     for item in row
        #     if item
        # ]

        data_rows = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]

        non_empty_data_rows = clean_rows(data_rows)

        engine = create_engine(
            "postgresql://postgres:local_password@localhost:5432/postgres"  # pragma: allowlist secret
        )
        Base.metadata.create_all(engine)

        # Delete all pre-existing records
        with Session(engine) as session, session.begin():
            session.query(AllBudgetCurrent).delete()

        # create session and add objects
        with Session(engine) as session, session.begin():
            importer = SQLTableImporter(session)
            adapter = SQLTableImportAdapter(AllBudgetCurrent)
            adapter.column_names = [
                "CAN",
                "Sys_Budget_ID",
                "Project_Title",
                "CIG_Name",
                "CIG_Type",
                "Line_Desc",
                "Date_Needed",
                "Amount",
                "PSCFee_Amount",
                "Status",
                "Comments",
                "Total_Bud_Cur",
                "All_Bud_Prev",
                "Delta",
            ]
            importer.append(adapter)
            save_data(
                importer, {adapter.get_name(): non_empty_data_rows}, file_type=DB_SQL
            )

        # with Session(engine) as session, session.begin():
        #     query_sets = session.query(AllBudgetCurrent).all()
        #     reader = QuerysetsReader(query_sets, header_row)
        #     results = reader.to_array()
        #     import json
        #
        #     json.dumps(list(results))


if __name__ == "__main__":
    luigi.build([ExtractExcelLoadTable()], local_scheduler=True)
